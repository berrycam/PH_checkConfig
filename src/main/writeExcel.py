#!/usr/bin/env python
# -*- coding:utf-8 -*-
# @Time     :2018/8/27
# @Author   :chenaimei
# @Function :将列表写入EXCEL

import logging
import openpyxl
from src.main.readConfig import ReadConfig
"""
说明：不使用xlsxwriter改使用openpyxl
原因：xlsxwriter不可excel进行编辑，导致出错
"""

class WriteExcel():
    def __init__(self):
        self.path = ReadConfig().get_file_path("result_path")

    def write2excel(self, arg_fileName, java_list, json_list, same_list):
        """
        :param arg_fileName: 进行字段对比的文件名，如A
        :param java_list: 该文件名A所对应java文件中的字段
        :param json_list: 该文件名A所对应json配置中的字段
        :param json_list: 该文件名A中，java与json均包含的字段
        :将上述arg_fileName作为excel的文件名，字段写入excel列
        """
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        # 设置列宽与行高
        worksheet.column_dimensions['A'].width = 40.0
        worksheet.column_dimensions['B'].width = 40.0
        worksheet.column_dimensions['C'].width = 40.0
        # 设置列宽、写入表头
        worksheet.title="字段对比结果"
        worksheet['A1'] = '仅JAVA端有：'
        worksheet['B1'] = '仅WEB端有：'
        worksheet['C1'] = 'JAVA端与WEB端均包含：'
        # 将java_list 写入Excel的第1列
        for i in range(len(java_list)):
            worksheet.cell(row=i+2, column=1, value=java_list[i])
        print(arg_fileName, "表----java_list：", len(java_list))

        # 将json_list 写入Excel的第2列
        for j in range(len(json_list)):
            worksheet.cell(row=j+2, column=2, value=json_list[j])
        print(arg_fileName, "表---json_list：", len(json_list))

        # 将same_list 写入Excel的第3列
        for k in range(len(same_list)):
            worksheet.cell(row=k+2, column=3, value=same_list[k])
        print(arg_fileName, "表---json_list：", len(same_list))

        tmp_name = self.path + '\\' + arg_fileName + ".xlsx"
        workbook.save(tmp_name)
        print(arg_fileName + " : 对比结果已写入Excel ! ")
        logging.debug("%s:对比结果已写入excel", arg_fileName)






